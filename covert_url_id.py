import openpyxl
from openpyxl import Workbook
import time
from selenium import webdriver
list_id=[]
def get_id(list_link):
    time.sleep(20)
    for link in list_link:
        driver = webdriver.Chrome(executable_path="/path_file_drive") # path drive chrome
        driver.get("https://lookup-id.com")
        a = driver.find_element_by_id("input_url")
        a.send_keys(link)
        b = driver.find_element_by_id("facebook_lookup_botton")
        b.click()
        try:
            id_=driver.find_element_by_id("code").text
            print([link,id_])
            list_id.append([link,id_])
            driver.close()
        except:
            print(link+' error')
            driver.close()
    return list_id
def import_file(path_file):

	list_link = [] #creat list link facebook 
	file_xlsx = openpyxl.load_workbook(path_file)
	file_sheet = file_xlsx.active #load file xlsx
	for row_ in range(file_sheet.max_row + 1): #read all row 
		value = file_sheet.cell(row = row_ + 1 ,column = 1).value #get value from row
		list_link.append(value)
	return list_link
def export_data(path_file_export,path_file_import):
    export_file = Workbook()
    sheet = export_file.active
    for list_ in get_id(import_file(path_file_import)):
        sheet.append(list_)
    export_file.save(path_file_export)






