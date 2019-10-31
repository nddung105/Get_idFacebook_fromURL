from selenium import webdriver
import time
import openpyxl
from openpyxl import Workbook
drive = webdriver.Chrome(executable_path="/yourPathDrive")

def get_id(email,pass_,list_link):
    list_id=[]
    drive.get("https://www.facebook.com/")
    send_email = drive.find_element_by_id("email")
    send_email.send_keys(email)
    send_pass = drive.find_element_by_id("pass")
    send_pass.send_keys(pass_)
    login = drive.find_element_by_id("loginbutton").click()
    for link in list_link:
        time.sleep(10)
        drive.get(link)
        id_facebook = drive.find_element_by_class_name("photoContainer").find_elements_by_tag_name("a")[0].get_attribute("href").split('=')[-1]
        list_id.append([link,id_facebook])
    return list_id

def import_file(path_file):

	list_link = [] #creat list link facebook 
	file_xlsx = openpyxl.load_workbook(path_file)
	file_sheet = file_xlsx.active #load file xlsx
	for row_ in range(file_sheet.max_row + 1): #read all row 
		value = file_sheet.cell(row = row_ + 1 ,column = 1).value #get value from row
		list_link.append(value)
        
	return list_link

def export_data(email,pass_,path_file_export,path_file_import):
    export_file = Workbook()
    sheet = export_file.active
    for list_ in get_id(email,pass_,import_file(path_file_import)):
        sheet.append(list_)
    export_file.save(path_file_export)


